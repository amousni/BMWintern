# -*- coding: utf-8 -*-
"""
Created on Thu May 23 13:24:48 2019

@author: qxp7153
"""

import pandas as pd
from openpyxl import load_workbook

#读取人名信息
def load_name_list():
    df = pd.read_csv('name_list.csv').set_index('name')
    return [list(df.index), df.T.to_dict()]

#department_dic = {'You Dragon':'TT','Lu Sandy':'PT', 'Li Gang':'EE','Yan Joshua':'PT','Song Colin':'DT', 'Yang Bo':'DT', 'Gao Yongyuan':'DT','Song Min':'DT', 'Li Zhaojun':'DT', 'Lu kejie':'DT', 'Geng Charlie':'DT', 'Tan Lunan':'DT', 'Jin Mingjun':'DT', 'Liu Luno':'DT', 'Wei Yunqian':'EE', 'Fu Rich':'EE', 'Meng Fan Bo':'EE', 'Ren Fei':'EE', 'You Marshal':'EE', 'Wang David':'EE', 'Li Feixue':'EE', 'Qu Zuoqi':'EE', 'Zhou Ke':'EE', 'Zhang Yuan':'PT', 'Shi Wei':'PT', 'Feng Jianquan':'PT', 'Wang Frank':'PT', 'Huang Sean':'PT', 'Ma Ben':'PT', 'Shu Yongcheng':'PT', 'Zhao Shelina':'TT', 'Zheng Annamaria':'TT','Chu Zhaoxian':'TT'}

#处理escalation email
def escalation_email(week_number):
    
    department_dic = load_name_list()[-1]
    
    #打开escalation
    file_name = 'Escalation Email Weekly Report CW' + week_number + '.xlsx'
    
    #打开reply sheet
    reply_sheet = 'CW' + week_number + ' Reply'
    
    #reply sheet dataframe
    dframe = pd.read_excel(file_name, sheet_name=reply_sheet)
    
    #reply个数
    row = dframe.shape[0]
    
    #需要的reply information
    reply = {'Name':[], 'Department':[], 'Title':[], 'Reply Time':[]}
    for i in range(row):
        name = dframe.loc[i,'From'].split(',')[0]
        reply['Name'].append(name)
        reply['Department'].append(department_dic[name]['department'])
        reply['Title'].append(dframe.loc[i, 'Subject'].split(':')[-1])
        reply_time_str = str(dframe.loc[i, 'Received']).split(' ')[-1]
        h = int(reply_time_str.split(':')[0])
        m = int(reply_time_str.split(':')[1])
        reply_time = h*60 + m
        reply['Reply Time'].append(reply_time)
    
    #打开email sheet
    email_sheet = 'CW' + week_number + ' Email'
    dframe = pd.read_excel(file_name, sheet_name=email_sheet)
    row = dframe.shape[0]
    
    #email information
    email = {'Title':[], 'Submission Time':[], 'isWeekday':[], 'isOuttime':[]}
    for i in range(row):
        email['Title'].append(dframe.loc[i, 'Subject'].split(':')[-1])
        submission_time_str = str(dframe.loc[i, 'Received']).split(' ')[-1]
        d = str(dframe.loc[i, 'Received']).split(' ')[0]
        if d == 'Sat' or d == 'Sun':
            email['isWeekday'].append(0)
        else:
            email['isWeekday'].append(1)
        h = int(submission_time_str.split(':')[0])
        m = int(submission_time_str.split(':')[1])
        submission_time = h*60 + m
        email['Submission Time'].append(submission_time)
        if submission_time <= 510 or submission_time >= 1050:
            email['isOuttime'].append(1)
        else:
            email['isOuttime'].append(0)
    
    #按照Title列合并reply和email dataframe，outer保证所有数据被录入        
    reply_df = pd.DataFrame(reply)
    email_df = pd.DataFrame(email)
    df = pd.merge(reply_df,email_df,on='Title',how='right')
    tf = df.isnull()
    response_time = []
    less_30 = []
    for i in range(df.shape[0]):
        if not tf.loc[i, 'Reply Time'] and not tf.loc[i, 'Submission Time']:
            subtime = df.loc[i, 'Reply Time'] - df.loc[i, 'Submission Time']
            if subtime < 0:
                subtime += 1440
            response_time.append(subtime)
            if subtime >= 0 and subtime <=30:
                less_30.append(1)
            else:
                less_30.append(0)
        else:
            response_time.append(-1)
            less_30.append(-1)
    df['Response Time'] = response_time
    df['No More Than 30'] = less_30
    data = df.drop_duplicates(subset=['Title','Submission Time'],keep='first',inplace=False)
    escalation = 'Escalation Analyzing CW' + week_number + '.xlsx'
    data.to_excel(escalation,encoding="utf_8_sig")
    
    #写入vlookup
    wb = load_workbook(escalation)
    ws = wb.active
    ws['L2'] = 'PT'
    ws['L3'] = 'DT'
    ws['L4'] = 'EE'
    ws['L5'] = 'TT'
    ws['M1'] = 'WORKDAY'
    ws['N1'] = 'WEENKEND'
    ws['O1'] = 'AVERAGE'
    ws['Q1'] = 'response time<30'
    ws['M2'] = '=COUNTIFS(C:C,"PT",G:G,"1")'
    ws['M3'] = '=COUNTIFS(C:C,"DT",G:G,"1")'
    ws['M4'] = '=COUNTIFS(C:C,"EE",G:G,"1")'
    ws['M5'] = '=COUNTIFS(C:C,"TT",G:G,"1")'
    ws['N2'] = '=COUNTIFS(C:C,"PT",G:G,"0")'
    ws['N3'] = '=COUNTIFS(C:C,"DT",G:G,"0")'
    ws['N4'] = '=COUNTIFS(C:C,"EE",G:G,"0")'
    ws['N5'] = '=COUNTIFS(C:C,"TT",G:G,"0")'
    ws['O2'] = '=AVERAGEIFS(I:I,G:G,"1",C:C,"PT",H:H,"0")'
    ws['O3'] = '=AVERAGEIFS(I:I,G:G,"1",C:C,"DT",H:H,"0")'
    ws['O4'] = '=AVERAGEIFS(I:I,G:G,"1",C:C,"EE",H:H,"0")'
    ws['O5'] = '=AVERAGEIFS(I:I,G:G,"1",C:C,"TT",H:H,"0")'
    ws['Q2'] = '=COUNTIFS(J:J,"1",C:C,"PT")'
    ws['Q3'] = '=COUNTIFS(J:J,"1",C:C,"DT")'
    ws['Q4'] = '=COUNTIFS(J:J,"1",C:C,"EE")'
    ws['Q5'] = '=COUNTIFS(J:J,"1",C:C,"TT")'
    ws['L8'] = 'Out of service time'
    ws['M8'] = '=COUNTIF(H:H,"1")'
    wb.save(escalation)
    print('-'*50)
    print('Finish escalation analyzing!')
    
def main():
    week_number = input('input week number:')
    escalation_email(week_number)
    
if __name__ == '__main__':
    main()