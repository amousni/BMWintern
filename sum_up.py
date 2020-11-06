# -*- coding: utf-8 -*-
"""
Created on Thu Jun 13 17:07:28 2019

@author: amousni
"""

"""
Update log

-----1.1-----
    1.更新xls2xlsx函数
        1.1.增加解压zip功能
        1.2.增加合并xlsx功能
    2.合并所有py文件为sum_up并计划打包exe
    3.更新weekly report guide文件
    
-----1.2-----
    1.编写load_name_list函数并储存本地name_list.csv文件
    2.替换所有函数内部的name_list以及department_dic信息并更新为load_name_list函数
    3.更新weekly report guide文件
    
"""

import os
import os.path
import win32com.client as win32
import time
import zipfile
import pandas as pd
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import datetime
import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)

#读取人名以及department信息

#---------------------！！！需要特殊注意！！！-------------------------
#   在name_list.csv中读取出的dataframe后
#   由于dataframe中可能出现多个columns
#   因此在dataframe转字典的过程中，key对应的value为字典，value字典中以column为key
#   因此需要加一个key值，即['department']以获取人名对应的分组信息
#---------------------！！！需要特殊注意！！！-------------------------
def load_name_list():
    df = pd.read_csv('name_list.csv').set_index('name')
    return [list(df.index), df.T.to_dict()]

#extract zip files from PuMA
#xls to xlsx
#concat the xlsx and save as open/created/modified
def xx(week_number):

    #桌面路径
    root = 'C:\\Users\\qxp7153\\Desktop\\'   
    l = ['open', 'modified', 'created']
    print('-'*50)
    print("MAKE SURE ALL ZIP FILES ARE DOWNLOADED AND PUT IN RIGHT FILEFOLDERS!")
    print('-'*50)
    
    #解压缩包
    for i in l:
        rootdir = root + i
        for parent, dirnames, filenames in os.walk(rootdir):
            for fn in filenames:
                filedir = os.path.join(parent, fn)
                if filedir.endswith('.zip'):
                    with zipfile.ZipFile(filedir) as zfile:
                        zfile.extractall(path = parent)
    print("zip files were extracted!")
    print('-'*50)
    print("xls files are changing to xlsx!")
    print('-'*50)
    print('44.8s needed for 10MB xls document')
    print('-'*50)
    
    #xls转xlsx
    for i in l:
        rootdir = root + i
        for parent, dirnames, filenames in os.walk(rootdir):
            for fn in filenames:
                starttime = time.time()
                filedir = os.path.join(parent, fn)
                if filedir.endswith('.xls'):
                    print(filedir)
                    excel = win32.gencache.EnsureDispatch('Excel.Application') 
                    wb = excel.Workbooks.Open(filedir) # xlsx: FileFormat=51 # xls: FileFormat=56,
                    wb.SaveAs(filedir.replace('xls', 'xlsx'), FileFormat=51)
                    wb.Close()
                    excel.Application.Quit()
                    endtime = time.time()
                    print(endtime-starttime)
                    print('-'*50)
    print('Finish all xls to xlsx!')
    print('-'*50)
    print("xlsx files are under concat!")
    print('-'*50)
    
    #xlsx合并   
    for i in l:
        file_name = 'CW' + week_number + ' ' + i + '.xlsx'
        a = {}
        df = pd.DataFrame(a) 
        rootdir = root + i
        for parent, dirnames, filenames in os.walk(rootdir):
            for fn in filenames:
                filedir = os.path.join(parent, fn)
                if filedir.endswith('.xlsx'):
                    temp_df = pd.read_excel(filedir)
                    temp_df = temp_df.iloc[:, 0:25]
                    df = pd.concat([df, temp_df], axis=0)
        df.to_excel(file_name)
        
        #打印合并后的数据个数，以确定文件位置以及文件合并是否正确
        data_no = df.shape[0]
        print("%s is saved!"%file_name)
        print("The number of data in %s is %d"%(file_name,data_no))
        print("-"*50)
        
#需要匹配的人名列表，modified文件中C/F/R列title
modified_index = ['Case no.', 'VIN no. (last 7 digits)', 'Previous recommendations/queries/additional information']

#匹配recommendation与additional information中的回复人信息与回复次数
#需要注意的是：
#   1.additional information中会有一部分回复信息
#   2.recommendation与additional information中有回复时间，仅统计半年内的回复信息
#   3.recommendation的回复人一定是组内的人，additional information中不一定
def recom_pair(recom_str):
    
    #name_list = ['Kun Xu', 'Annamaria Zheng', 'xing bin', 'bin xing', 'Chunhu Zhang', 'Don Wang', 'Weiyu Hou', 'Kejie Lu', 'Wenjie Li', 'Hailong Zhao', 'Qi Yang', 'Yongyuan Gao', 'Ruoyun Wang', 'Frank Wang', 'Jian Xu', 'Min Song', 'Ben Ma', 'Jianquan Feng', 'Hu Sun', 'Joshua Yan', 'Raul Bernal', 'Yali Jiang', 'Kai Wang', 'Chengyan Sun', 'Hong Wang', 'Xin Tian', 'Xin Tian', 'Alice Yang', 'Fady Bawatna', 'Li Wang', 'Peter Chen', 'Angela Feng', 'Pengcheng Gao', 'lei Zhang', 'Dongli Liu', 'Lesley lan', 'Zhenhua Li', 'Alexander Lang', 'Zhi Zhang', 'Shuai Ni', 'Donghui Zhang', 'Rui Miao', 'Benzy Qin', 'Ping Hao', 'TB GZ', 'hongxing zhao', 'Rocky Chen', 'Becky Zhang', 'yuan zhou3', 'juan zhang', 'Xin Tian', 'Frank Zheng', 'Ivan Wang', 'Qiang Zhong', 'Guoxiang Deng', 'Alan Zhang', 'Test1 Puma', 'Reader Power User', 'Peter Loertzing', 'Carsten Gruetzmacher', 'Christoph Brettle', 'Tao Zhao', 'Wendy Wu', 'Baolin Xiang', 'Bin Fu', 'Joanna Zhong', 'Mahathi Murthy', 'Rishika Menon', 'Yali Jiang', 'Xiaomao Wu', 'Ting Xu', 'Yali Jiang', 'Hong Wang', 'ruiqing Zhang', 'Yang Li', 'David Li', 'Yu Hou', 'Jing Dai', 'Joanna Zhong', 'Hong Wang', 'Fengjie Lu', 'Jie Tan', 'Sean Huang', 'Dan Wu', 'Fan Bo Meng', 'Anna Yang', 'Fernando Ferrer', 'Rocky Chen', 'Becky Zhang', 'Fei Ren', 'yuan zhou3', 'juan zhang', 'Jie Yang', 'Xin Tian', 'Xin Tian', 'Jintu Yang', 'Andrew Fletcher', 'Yunqian Wei', 'Xiaoxi Liu', 'Alice Yang', 'Bo Yang', 'Fady Bawatna', 'Huifeng Cong', 'Li Wang', 'Peter Chen', 'Marshal You', 'Zuoqi Qu', 'Graham Boyd', 'Jamie Zhang', 'Angela Feng', 'Yanjun Gu', 'Pengcheng Gao', 'Fangchao Xu', 'Pengcheng Gao', 'Yan Yang', 'Lei Li', 'Leo Li', 'Juan Zhang', 'Shixin Dong', 'juan zhang', 'Zhaojun Li', 'Jie Tan', 'David Wang', 'Marc Ma', 'Dawei Wang', 'Dongsheng Liu ', 'Zhengyun Lu', 'Lipeng Wang', 'Lei Liu', 'Yuping li', 'lei Zhang', 'Quan Huang', 'Christian Schoenmeyer', 'Maple Huo', 'Laney Zhang', 'Joachim Fluegge', 'Gordon Zhao', 'Chris Dong', 'Artem Lobanow', 'Zhe Yang', 'Gary Tian', 'Zhicong Xie', 'Dongli Liu', 'Lesley lan', 'Steven Miao', 'Alex Li', 'Zhenhua Li', 'Hao Ling', 'Ao Shen', 'Frank Brandel', 'Mareike Volkmer', 'Arpad Toth', 'Xuqiang Yin', 'Qiang Zhong', 'Andreas Sedlak', 'Ting Jin', 'Zhihai Wu', 'Chun Zou', 'Yuan Zhang', 'Collie Wang', 'Yiding Zhang', 'Guoxiang Deng', 'Season Liu', 'Jonathan Zhu', 'Billy Xiao', 'Yin Li', 'Alan Zhang', 'Ke Chen', 'Justin Jiang', 'Alexander Lang', 'Weiming Hou', 'Gabrielle Ge', 'Billy Fang', 'Wilson Xu', 'Dawei Zhang', 'Zhi Zhang', 'Shuai Ni', 'Joao Martins', 'Wei Shi', 'Gavin Cai', 'dan zhang', 'Justin Zeng', 'Jimmy Ji', 'Spike Zhao', 'Juliet Zhang', 'Jiangtao Chi', 'Steven Jiang', 'Weiliang Huang', 'Ruifeng Liu', 'Ming Liu', 'Yan Ke', 'Vicente Liu', 'Chris Liu', 'Mandy Du', 'Fred Zhang', 'Qinran Luo', 'Harry Zhang', 'Ming Yuan', 'Benzy Qin', 'Jian Ma', 'Walter Gassner', 'Steven Wu', 'Galen Guo', 'Feal Wang', 'Colin Song', 'Fred Feng', 'Ping Hao', 'Jianmin Xin', 'Jarod Yang', 'TB GZ', 'Tony Xu', 'Mingjia Ma', 'Tianshu Li', 'Rich Fu', 'Wayne Zhang', 'Alexander Lang', 'hongxing zhao', 'Christian Kluth', 'Friedrich Mursch', 'Alexander Fischer', 'John White', 'Richard Brown', 'Test1 Puma', 'Reader Power User', 'Thomas Riedl', 'Mark Werner', 'Harald Kammerl', 'Jean Claude Chauveau', 'Frank Schoenrock', 'Stefan Rosse', 'Horst Redner', 'Regan Chen', 'Friedrich Mursch', 'Ian Weston', 'Andreas Herkert', 'Sebastien Beck', 'Thomas Drexler', 'Markus Gamisch', 'Wenjin Chen', 'Richard Wimmer', 'Huashan Wang', 'Yingxue Peng', 'Stefan Hofmann', 'Bob White', 'Christian Mueck', 'Wendi Yi', 'Feng Fred', 'Di Lai', 'Armin Meyer', 'Maoqi Cui', 'Hebe He', 'zhihai wu', 'Sinan Oeztan', 'Yingxue Peng', 'Rex Jiang', 'Frank Zintl', 'Sherry Lu', 'Junli Liang', 'Tim Shi', 'Yuhuo Ma', 'lilong zhao', 'Allen Mao', 'Hailong Li', 'Shawn Chen', 'Leiming He', 'Ermera Setyadi', 'Kimreol Chen', 'Stefan Krott', 'Xiaohui Li', 'Liping Hong', 'Jay Lee', 'Christian Klenovsky', 'Feixue Li', 'Jeffrey Jiang', 'Shelina Zhao', 'Zhang Zhe', 'Jian Guo', 'weiqiang mu', 'Sabrina Fischer', 'Zhi Cui', 'Ryan Qian', 'Zenan Wen', 'Yinbo Shen', 'Zhaoxian Chu', 'Dragon You', 'Wenhao Li', 'Zutao Kou', 'Lunan Tan', 'Charlie Geng', 'Sandy Lu', 'Ke Zhou', 'Haoxin Liu', 'Gang Li', 'Mingjun Jin', 'Levi Yang', 'Junyan Liu', 'Yu Wang', 'Zhiqiang Chen', 'Yongcheng Shu', 'Shengnan Si', 'Lei Yang ', 'Bin Xing', 'Hanan Wu', 'Jack Zhang']
    #pingpong中正则匹配的人名是名+姓，name_list中的是姓+名，需要调换一下顺序
    unfinished_name_list = load_name_list()[0]
    name_list = []
    for i in unfinished_name_list:
        last_first = i.split(' ')
        name = last_first[-1] + ' ' + last_first[0]
        name_list.append(name)
    
    #时间节点，当天至半年前
    end_day = datetime.date.today()
    start_day = end_day - datetime.timedelta(days=183)
    
    #正则匹配回复时间与回复者人名
    pattern = re.compile('Recommendation   (.*?)   .*?   (.*? .*?) ')
    
    #匹配信息，一个元组列表，元组中第一个元素为回复时间，第二个元素为回复人名
    s = pattern.findall(recom_str)
    
    #人名列表
    recom_name_list = []
    
    #时间匹配计数
    flag = 0
    for i in s:
        mdy_str = i[0].split('/')
        mdy = datetime.date(int(str(20) + mdy_str[2]),int(mdy_str[0]),int(mdy_str[1]))
        if mdy <= end_day and mdy >= start_day:
            recom_name_list.append(i[1])
            flag += 1
            
    #Additional information正则匹配
    pattern_a = re.compile('Additional information   (.*?)   .*?   (.*? .*?) ')
    s_a = pattern_a.findall(recom_str)
    for i in s_a:
        mdy_str = i[0].split('/')
        mdy = datetime.date(int(str(20) + mdy_str[2]),int(mdy_str[0]),int(mdy_str[1]))
        if mdy <= end_day and mdy >= start_day:
            if i[1] in name_list:
                recom_name_list.append(i[1])
                flag += 1
                
    #人名合并
    name = ''
    recom_name = set(recom_name_list)
    recom_name_list = list(recom_name)
    for i in range(len(recom_name_list)):
        if i == 0:
            name = name + recom_name_list[i]
        else:
            name = name + '|' + recom_name_list[i]
            
    #返回匹配个数与人名信息
    return [flag, name]
        
#Pingpong分析
def pp(week_number):

    #读取本周modified文件
    #week_number为本周周报号, file_name表示本周modified文件名
    file_name = 'CW' + week_number + ' MODIFIED' + '.xlsx'

	#wb为本周modified文件
    wb = xlrd.open_workbook(filename = file_name)

	#打开wb中sheet1
    sheet1 = wb.sheet_by_index(0)
    cla = sheet1.row_values(0)

	#sheet1中的C/F/R列
    case_no = cla.index(modified_index[0])
    vin = cla.index(modified_index[1])
    recom = cla.index(modified_index[2])
    #sheet1中列存为列表方便写入新文件
    sheet_index_list = [case_no, vin, recom]
    data_number = len(sheet1.col_values(case_no))

    #记录数据
    #case_no_value = sheet1.col_values(case_no)[0:]
    #vin_value = sheet1.col_values(vin)[0:]
    recom_value = sheet1.col_values(recom)[0:]

    #将C/F/R列写入新文件
    pingpong_file_name = 'Ping-Pong Analyzing CW' + week_number + '.xlsx'
    wb_pingpong = Workbook()
    sheet = wb_pingpong.active
    sheet.title = 'Statisitc'
    #三列，储存在sheet_index_list中
    #flag读写行数标注
    #flag = 1
    for row in range(data_number):
        #第1行记录表头信息
        if row == 0:
            sheet.cell(row+1,1,sheet1.col_values(sheet_index_list[0])[row])
            sheet.cell(row+1,2,sheet1.col_values(sheet_index_list[1])[row])
            sheet.cell(row+1,5,sheet1.col_values(sheet_index_list[2])[row])
            sheet.cell(row+1,4,'Recommend Person')
            sheet.cell(row+1,3,'Numbers OF reply')
        else:
            #记录前三列C/F/R列信息
            sheet.cell(row+1,1,sheet1.col_values(sheet_index_list[0])[row])
            sheet.cell(row+1,2,sheet1.col_values(sheet_index_list[1])[row])
            sheet.cell(row+1,5,sheet1.col_values(sheet_index_list[2])[row])
            #recommendation_string
            recom_str = recom_value[row]                
            pair_info = recom_pair(recom_str)
            sheet.cell(row+1,4,pair_info[1])
            sheet.cell(row+1,3,pair_info[0])

    #保存Pingpong文件，xlsx格式保存
    wb_pingpong.save(pingpong_file_name)
    print('-'*10)
    print('Finish Pingpong analyzing!')
    
#open case的recommendation信息匹配
def recom_pair_for_opencase(recom_str):
    
    #name_list = ['guoqing qian','Lei Yang','Dongsheng Liu','yuxian hu','Hailong Yu','Kun Xu', 'Annamaria Zheng', 'xing bin', 'bin xing', 'Chunhu Zhang', 'Don Wang', 'Weiyu Hou', 'Kejie Lu', 'Wenjie Li', 'Hailong Zhao', 'Qi Yang', 'Yongyuan Gao', 'Ruoyun Wang', 'Frank Wang', 'Jian Xu', 'Min Song', 'Ben Ma', 'Jianquan Feng', 'Hu Sun', 'Joshua Yan', 'Raul Bernal', 'Yali Jiang', 'Kai Wang', 'Chengyan Sun', 'Xin Tian', 'Xin Tian', 'Alice Yang', 'Fady Bawatna', 'Li Wang', 'Peter Chen', 'Angela Feng', 'Pengcheng Gao', 'lei Zhang', 'Dongli Liu', 'Lesley lan', 'Zhenhua Li', 'Alexander Lang', 'Zhi Zhang', 'Shuai Ni', 'Donghui Zhang', 'Rui Miao', 'Benzy Qin', 'Ping Hao', 'TB GZ', 'hongxing zhao', 'Rocky Chen', 'Becky Zhang', 'yuan zhou3', 'juan zhang', 'Xin Tian', 'Frank Zheng', 'Ivan Wang', 'Qiang Zhong', 'Guoxiang Deng', 'Alan Zhang', 'Test1 Puma', 'Reader Power User', 'Peter Loertzing', 'Carsten Gruetzmacher', 'Christoph Brettle', 'Tao Zhao', 'Wendy Wu', 'Baolin Xiang', 'Bin Fu', 'Joanna Zhong', 'Mahathi Murthy', 'Rishika Menon', 'Yali Jiang', 'Xiaomao Wu', 'Ting Xu', 'Yali Jiang', 'ruiqing Zhang', 'Yang Li', 'David Li', 'Yu Hou', 'Jing Dai', 'Joanna Zhong', 'Fengjie Lu', 'Jie Tan', 'Sean Huang', 'Dan Wu', 'Fan Bo Meng', 'Anna Yang', 'Fernando Ferrer', 'Rocky Chen', 'Becky Zhang', 'Fei Ren', 'yuan zhou3', 'juan zhang', 'Jie Yang', 'Xin Tian', 'Xin Tian', 'Jintu Yang', 'Andrew Fletcher', 'Yunqian Wei', 'Xiaoxi Liu', 'Alice Yang', 'Bo Yang', 'Fady Bawatna', 'Huifeng Cong', 'Li Wang', 'Peter Chen', 'Marshal You', 'Zuoqi Qu', 'Graham Boyd', 'Jamie Zhang', 'Angela Feng', 'Yanjun Gu', 'Pengcheng Gao', 'Fangchao Xu', 'Pengcheng Gao', 'Yan Yang', 'Lei Li', 'Leo Li', 'Juan Zhang', 'Shixin Dong', 'juan zhang', 'Zhaojun Li', 'Jie Tan', 'David Wang', 'Marc Ma', 'Dawei Wang', 'Dongsheng Liu ', 'Zhengyun Lu', 'Lipeng Wang', 'Lei Liu', 'Yuping li', 'lei Zhang', 'Quan Huang', 'Christian Schoenmeyer', 'Maple Huo', 'Laney Zhang', 'Joachim Fluegge', 'Gordon Zhao', 'Chris Dong', 'Artem Lobanow', 'Zhe Yang', 'Gary Tian', 'Zhicong Xie', 'Dongli Liu', 'Lesley lan', 'Steven Miao', 'Alex Li', 'Zhenhua Li', 'Hao Ling', 'Ao Shen', 'Frank Brandel', 'Mareike Volkmer', 'Arpad Toth', 'Xuqiang Yin', 'Qiang Zhong', 'Andreas Sedlak', 'Ting Jin', 'Zhihai Wu', 'Chun Zou', 'Yuan Zhang', 'Collie Wang', 'Yiding Zhang', 'Guoxiang Deng', 'Season Liu', 'Jonathan Zhu', 'Billy Xiao', 'Yin Li', 'Alan Zhang', 'Ke Chen', 'Justin Jiang', 'Alexander Lang', 'Weiming Hou', 'Gabrielle Ge', 'Billy Fang', 'Wilson Xu', 'Dawei Zhang', 'Zhi Zhang', 'Shuai Ni', 'Joao Martins', 'Wei Shi', 'Gavin Cai', 'dan zhang', 'Justin Zeng', 'Jimmy Ji', 'Spike Zhao', 'Juliet Zhang', 'Jiangtao Chi', 'Steven Jiang', 'Weiliang Huang', 'Ruifeng Liu', 'Ming Liu', 'Yan Ke', 'Vicente Liu', 'Chris Liu', 'Mandy Du', 'Fred Zhang', 'Qinran Luo', 'Harry Zhang', 'Ming Yuan', 'Benzy Qin', 'Jian Ma', 'Walter Gassner', 'Galen Guo', 'Feal Wang', 'Colin Song', 'Fred Feng', 'Ping Hao', 'Jianmin Xin', 'Jarod Yang', 'TB GZ', 'Tony Xu', 'Mingjia Ma', 'Tianshu Li', 'Rich Fu', 'Wayne Zhang', 'Alexander Lang', 'hongxing zhao', 'Christian Kluth', 'Friedrich Mursch', 'Alexander Fischer', 'John White', 'Richard Brown', 'Test1 Puma', 'Reader Power User', 'Thomas Riedl', 'Mark Werner', 'Harald Kammerl', 'Jean Claude Chauveau', 'Frank Schoenrock', 'Stefan Rosse', 'Horst Redner', 'Regan Chen', 'Friedrich Mursch', 'Ian Weston', 'Andreas Herkert', 'Sebastien Beck', 'Thomas Drexler', 'Markus Gamisch', 'Wenjin Chen', 'Richard Wimmer', 'Huashan Wang', 'Yingxue Peng', 'Stefan Hofmann', 'Bob White', 'Christian Mueck', 'Wendi Yi', 'Feng Fred', 'Di Lai', 'Armin Meyer', 'Maoqi Cui', 'Hebe He', 'zhihai wu', 'Sinan Oeztan', 'Yingxue Peng', 'Rex Jiang', 'Frank Zintl', 'Sherry Lu', 'Junli Liang', 'Tim Shi', 'Yuhuo Ma', 'lilong zhao', 'Allen Mao', 'Hailong Li', 'Shawn Chen', 'Leiming He', 'Ermera Setyadi', 'Kimreol Chen', 'Stefan Krott', 'Xiaohui Li', 'Liping Hong', 'Jay Lee', 'Christian Klenovsky', 'Feixue Li', 'Jeffrey Jiang', 'Shelina Zhao', 'Zhang Zhe', 'Jian Guo', 'weiqiang mu', 'Sabrina Fischer', 'Zhi Cui', 'Ryan Qian', 'Zenan Wen', 'Yinbo Shen', 'Zhaoxian Chu', 'Dragon You', 'Wenhao Li', 'Zutao Kou', 'Lunan Tan', 'Charlie Geng', 'Sandy Lu', 'Ke Zhou', 'Haoxin Liu', 'Gang Li', 'Mingjun Jin', 'Levi Yang', 'Junyan Liu', 'Yu Wang', 'Zhiqiang Chen', 'Yongcheng Shu', 'Shengnan Si', 'Lei Yang ', 'Bin Xing', 'Hanan Wu', 'Jack Zhang']
    #pingpong中正则匹配的人名是名+姓，name_list中的是姓+名，需要调换一下顺序
    unfinished_name_list = load_name_list()[0]
    name_list = []
    for i in unfinished_name_list:
        last_first = i.split(' ')
        name = last_first[-1] + ' ' + last_first[0]
        name_list.append(name)
    
    #时间节点，当天至半年前
    end_day = datetime.date.today()
    start_day = end_day - datetime.timedelta(days=365)
    
    #正则匹配回复时间与回复者人名
    pattern = re.compile('Recommendation   (.*?)   .*?   (.*? .*?) ')
    
    #匹配信息，一个元组列表，元组中第一个元素为回复时间，第二个元素为回复人名
    s = pattern.findall(recom_str)
    
    #人名列表
    recom_name_list = []
    
    #时间匹配计数
    flag = 0
    for i in s:
        mdy_str = i[0].split('/')
        mdy = datetime.date(int(str(20) + mdy_str[2]),int(mdy_str[0]),int(mdy_str[1]))
        if mdy <= end_day and mdy >= start_day:
            recom_name_list.append(i[1])
            flag += 1
            
    #Additional information正则匹配
    pattern_a = re.compile('Additional information   (.*?)   .*?   (.*? .*?) ')
    s_a = pattern_a.findall(recom_str)
    for i in s_a:
        mdy_str = i[0].split('/')
        mdy = datetime.date(int(str(20) + mdy_str[2]),int(mdy_str[0]),int(mdy_str[1]))
        if mdy <= end_day and mdy >= start_day:
            if i[1] in name_list:
                recom_name_list.append(i[1])
                flag += 1
                
    #正则匹配到的人名是名+姓，department字典中是姓+名，需要调换一下顺序
    #如果正则匹配到人名，则取正则中的第一个匹配对象和最后一个匹配对象，否则name取0
    name = []
    if len(recom_name_list)>0:
        name0list = recom_name_list[0].split(' ')
        name0 = name0list[-1] + ' ' + name0list[0]
        name.append(name0)
        name1list = recom_name_list[0].split(' ')
        name1 = name1list[-1] + ' ' + name1list[0]
        name.append(name1)        
    else:
        name = [0]
    
    #返回两个人名的意义在于，department字典中的人名可能不全
    #如果第一个不能在字典中匹配，则匹配最后一个，如果正则中只匹配到一个，则name中的两个人名相同
    #返回的name为一个列表，如果匹配到人名，则name的长度为2，如果未匹配到，则长度为1，元素为0            
    return name

#open case analyzing
def oc(week_number, end_day):
    print("Please wait for about 15.80s...")
    
    #department_dic = {'Yang Levi':'DT','Li Keats':'DT','Huang Jeffrey':'PT','Zeng Justin':'DT','qian guoqing':'0','Yang Lei':'DT','Zou Chun':'PT','Yang Jintu':'EE','Guo Galen':'DT','Liu Dongsheng':'PT','hu yuxian':'PT','Zhang Chunhu':'EE','Yu Hailong':'EV','You Dragon':'TT','Xu Wilson':'DT','Si Shengnan':'DT','Wu Hanan':'PT','Zhang Harry':'PT','Cong Huifeng':'EE','Kou Zutao':'DT','Li Wenhao':'DT','China Alpine':'EE','li Yuping':'PT','Li Yin':'DT','Wang Yu':'EE','Liu Junyan':'DT','Wu Steven':'PT','Huang Weiliang':'EE','Ma Marc':'DT','Chen Zhiqiang':'DT','Yang Lei':'DT','Zhang Wayne':'EE','Dong Shixin':'PT','Chen Shawn':'PT','Xiang Baolin':'EE','Lu Sandy':'PT','Jiang Steven':'EE','xing bin':'EE','Wen Zenan':'EE','Li Gang':'EE','Yan Joshua':'PT','Song Colin':'DT', 'Yang Bo':'DT', 'Gao Yongyuan':'DT','Song Min':'DT', 'Li Zhaojun':'DT', 'Lu kejie':'DT', 'Geng Charlie':'DT', 'Tan Lunan':'EE', 'Jin Mingjun':'DT', 'Liu Luno':'DT', 'Wei Yunqian':'EE', 'Fu Rich':'EE', 'Meng Fan Bo':'EE', 'Ren Fei':'EE', 'You Marshal':'EE', 'Wang David':'EE', 'Li Feixue':'EE', 'Qu Zuoqi':'EE', 'Zhou Ke':'EE', 'Zhang Yuan':'PT', 'Shi Wei':'PT', 'Feng Jianquan':'PT', 'Wang Frank':'PT', 'Huang Sean':'PT', 'Ma Ben':'PT', 'Shu Yongcheng':'PT', 'Zhao Shelina':'TT', 'Zheng Annamaria':'TT','Chu Zhaoxian':'TT'}
    department_dic = load_name_list()[-1]    
    
    #file_name为输出的open case结果
    #cognos为从1月1号到end_day的cognos数据
    #open_file为本周的open case
    file_name = 'Open Cases from 20190101-2019' + end_day + '.xlsx'
    cognos = 'Case and Vehicle Details open 20190101-2019' + end_day + '.xlsx'
    open_file = 'CW' + week_number + ' open.xlsx'
    
    #读取cognos数据中的case id和urgency两列
    cognos_df_whole = pd.read_excel(cognos)
    cognos_df = pd.DataFrame(cognos_df_whole, columns=['Case Id', 'Urgency'])
    
    #打开open case
    open_df_whole = pd.read_excel(open_file)
    
    #提取open case中case no.和recommendation两列
    #将recommendation列改名
    #将open_df中的Case no.改名为和Cognos中一样的Case Id，目的是将open_df与cognos_df按相同的列合并
    open_df = pd.DataFrame(open_df_whole, columns=['Case no.', 'Previous recommendations/queries/additional information', 'Remarks'])
    open_df = open_df.rename(columns={'Case no.':'Case Id', 'Previous recommendations/queries/additional information':'Recommendation'})
    
    #将open case中的空值数据填为0
    open_df = open_df.fillna(0)
    
    #open case数据个数
    row = open_df.shape[0]
    
    #department列表，也可以改进为open_df['depart'] = []
    depart = []
    
    #对open case数据进行循环
    for i in range(row):
        
        #判断是否由recommendation数据，没有depart添加0
        if open_df.loc[i, 'Recommendation'] != 0:
            
            #正则匹配recommendation中的回复者信息
            name = recom_pair_for_opencase(open_df.loc[i, 'Recommendation'])
            
            #匹配到人名则len(name)>1
            if len(name) > 1:
                
                #判断第一个匹配到的人名是否在department字典中
                #如果不在，则用最后一个匹配到的人名
                #如果两个人名都不在，则depart添加-1
                
                #！！！需要特殊注意！！！
                #在name_list.csv中读取出的dataframe后
                #由于dataframe中可能出现多个columns
                #因此在dataframe转字典的过程中，key对应的value为字典，value字典中以column为key
                #因此需要加一个key值，即['department']以获取人名对应的分组信息
                if name[0] in department_dic.keys():
                    depart.append(department_dic[name[0]]['department'])
                else:
                    if name[1] in department_dic.keys():
                        depart.append(department_dic[name[1]]['department'])
                    else:
                        depart.append('-1')
            else:
                depart.append('0')
        else:
            depart.append('0')
            
    #把depart赋值给open_df
    open_df['team'] = depart
    
    #筛选remarks不为B1的数据，remarks为B1表示不需要回复，需要剔除
    open_df = open_df[open_df.Remarks != 'B1']
    open_df = pd.DataFrame(open_df, columns=['Case Id', 'Recommendation', 'team'])
    open_df = open_df.rename(columns={'Urgency.':'urgency'})
    
    #将open_df与按Case Id合并
    df = pd.merge(open_df,cognos_df,on='Case Id',how='left')
    
    #df中出现空值则填0
    df = df.fillna(0)
    
    #保存文件
    df.to_excel(file_name)
    print('-'*50)
    print('Finish Open case analyzing!')

#处理escalation email
def ee(week_number):
    
    #department_dic = {'You Dragon':'TT','Lu Sandy':'PT', 'Li Gang':'EE','Yan Joshua':'PT','Song Colin':'DT', 'Yang Bo':'DT', 'Gao Yongyuan':'DT','Song Min':'DT', 'Li Zhaojun':'DT', 'Lu kejie':'DT', 'Geng Charlie':'DT', 'Tan Lunan':'EE', 'Jin Mingjun':'DT', 'Liu Luno':'DT', 'Wei Yunqian':'EE', 'Fu Rich':'EE', 'Meng Fan Bo':'EE', 'Ren Fei':'EE', 'You Marshal':'EE', 'Wang David':'EE', 'Li Feixue':'EE', 'Qu Zuoqi':'EE', 'Zhou Ke':'EE', 'Zhang Yuan':'PT', 'Shi Wei':'PT', 'Feng Jianquan':'PT', 'Wang Frank':'PT', 'Huang Sean':'PT', 'Ma Ben':'PT', 'Shu Yongcheng':'PT', 'Zhao Shelina':'TT', 'Zheng Annamaria':'TT','Chu Zhaoxian':'TT'}
    department_dic = load_name_list()[-1]
    
    #打开escalation
    file_name = 'Escalation Email Weekly Report CW' + week_number + '.xlsx'
    
    #打开reply sheet
    reply_sheet = 'CW' + week_number + ' Reply'
    
    #reply sheet dataframe
    dframe = pd.read_excel(file_name, sheet_name=reply_sheet)
    
    #reply个数
    row = dframe.shape[0]
    
    #需要的reply information
    reply = {'Name':[], 'Department':[], 'Title':[], 'Reply Time':[]}
    for i in range(row):
        name = dframe.loc[i,'From'].split(',')[0]
        reply['Name'].append(name)
        
        #！！！需要特殊注意！！！
        #在name_list.csv中读取出的dataframe后
        #由于dataframe中可能出现多个columns
        #因此在dataframe转字典的过程中，key对应的value为字典，value字典中以column为key
        #因此需要加一个key值，即['department']以获取人名对应的分组信息
        reply['Department'].append(department_dic[name]['department'])
        reply['Title'].append(dframe.loc[i, 'Subject'].split(':')[-1])
        reply_time_str = str(dframe.loc[i, 'Received']).split(' ')[-1]
        h = int(reply_time_str.split(':')[0])
        m = int(reply_time_str.split(':')[1])
        reply_time = h*60 + m
        reply['Reply Time'].append(reply_time)
    
    #打开email sheet
    email_sheet = 'CW' + week_number + ' Email'
    dframe = pd.read_excel(file_name, sheet_name=email_sheet)
    row = dframe.shape[0]
    
    #email information
    email = {'Title':[], 'Submission Time':[], 'isWeekday':[], 'isOuttime':[]}
    for i in range(row):
        email['Title'].append(dframe.loc[i, 'Subject'].split(':')[-1])
        submission_time_str = str(dframe.loc[i, 'Received']).split(' ')[-1]
        d = str(dframe.loc[i, 'Received']).split(' ')[0]
        if d == 'Sat' or d == 'Sun':
            email['isWeekday'].append(0)
        else:
            email['isWeekday'].append(1)
        h = int(submission_time_str.split(':')[0])
        m = int(submission_time_str.split(':')[1])
        submission_time = h*60 + m
        email['Submission Time'].append(submission_time)
        if submission_time <= 510 or submission_time >= 1050:
            email['isOuttime'].append(1)
        else:
            email['isOuttime'].append(0)
    
    #按照Title列合并reply和email dataframe，outer保证所有数据被录入        
    reply_df = pd.DataFrame(reply)
    email_df = pd.DataFrame(email)
    df = pd.merge(reply_df,email_df,on='Title',how='right')
    tf = df.isnull()
    response_time = []
    less_30 = []
    for i in range(df.shape[0]):
        #判断reply time以及submission time是否非空
        #非空正常计算
        #空值填-1
        if not tf.loc[i, 'Reply Time'] and not tf.loc[i, 'Submission Time']:
            subtime = df.loc[i, 'Reply Time'] - df.loc[i, 'Submission Time']
            if subtime < 0:
                subtime += 1440
            response_time.append(subtime)
            if subtime >= 0 and subtime <=30:
                less_30.append(1)
            else:
                less_30.append(0)
        else:
            response_time.append(-1)
            less_30.append(-1)
    df['Response Time'] = response_time
    df['No More Than 30'] = less_30
    data = df.drop_duplicates(subset=['Title','Submission Time'],keep='first',inplace=False)
    escalation = 'Escalation Analyzing CW' + week_number + '.xlsx'
    data.to_excel(escalation,encoding="utf_8_sig")
    
    #写入vlookup
    wb = load_workbook(escalation)
    ws = wb.active
    ws['L2'] = 'PT'
    ws['L3'] = 'DT'
    ws['L4'] = 'EE'
    ws['L5'] = 'TT'
    ws['M1'] = 'WORKDAY'
    ws['N1'] = 'WEENKEND'
    ws['O1'] = 'AVERAGE'
    ws['Q1'] = 'response time<30'
    ws['M2'] = '=COUNTIFS(C:C,"PT",G:G,"1")'
    ws['M3'] = '=COUNTIFS(C:C,"DT",G:G,"1")'
    ws['M4'] = '=COUNTIFS(C:C,"EE",G:G,"1")'
    ws['M5'] = '=COUNTIFS(C:C,"TT",G:G,"1")'
    ws['N2'] = '=COUNTIFS(C:C,"PT",G:G,"0")'
    ws['N3'] = '=COUNTIFS(C:C,"DT",G:G,"0")'
    ws['N4'] = '=COUNTIFS(C:C,"EE",G:G,"0")'
    ws['N5'] = '=COUNTIFS(C:C,"TT",G:G,"0")'
    ws['O2'] = '=AVERAGEIFS(I:I,G:G,"1",C:C,"PT",H:H,"0")'
    ws['O3'] = '=AVERAGEIFS(I:I,G:G,"1",C:C,"DT",H:H,"0")'
    ws['O4'] = '=AVERAGEIFS(I:I,G:G,"1",C:C,"EE",H:H,"0")'
    ws['O5'] = '=AVERAGEIFS(I:I,G:G,"1",C:C,"TT",H:H,"0")'
    ws['Q2'] = '=COUNTIFS(J:J,"1",C:C,"PT")'
    ws['Q3'] = '=COUNTIFS(J:J,"1",C:C,"DT")'
    ws['Q4'] = '=COUNTIFS(J:J,"1",C:C,"EE")'
    ws['Q5'] = '=COUNTIFS(J:J,"1",C:C,"TT")'
    ws['L8'] = 'Out of service time'
    ws['M8'] = '=COUNTIF(H:H,"1")'
    wb.save(escalation)
    print('-'*50)
    print('Finish escalation analyzing!')

#读取response time数据并生成failed list文件
def fl(week_number):
    
    #输入上周假期信息
    #number_of_holiday表示上周假期天数
    #holiday_no_list储存假期为星期几以便于筛选failed list
    number_of_holiday = int(input("Input the number of holidays in last week, 0 for none:"))
    holiday_no_list = []
    if number_of_holiday != 0:
        print('Input holiday in last week(eg: 0501):')
        for i in range(number_of_holiday):
            holiday_str = input('Holiday string:')
            holiday = datetime.date(2019, int(holiday_str[0:2]), int(holiday_str[2:4]))
            holiday_no = holiday.strftime("%w")
            holiday_no_list.append(holiday_no)
    print('-'*50)
    print("Program is reading Response time data...")
    print('-'*50)
    
    #打开response time文件
    #从summary sheet中取出tc/urgent/rr的相应列数据信息
    #从Run8sheet中提取时间信息
    file_name = 'Response time 2019-CW%s.xlsm'%(week_number,)
    summary_df = pd.read_excel(file_name, sheet_name = 'Summary')
    tc = pd.DataFrame(summary_df, columns=['Case no.', 'Team', 'Result.1', 'Urgency'])
    urgent = pd.DataFrame(summary_df, columns=['Case no.', 'Team', 'Result.1', 'Urgency'])
    rr = pd.DataFrame(summary_df, columns=['Case no.', 'Team', 'Result.3', 'Urgency'])
    whole_time_df = pd.read_excel(file_name, sheet_name = 'Run8')
    time_df = pd.DataFrame(whole_time_df, columns = ['Case No.', 'Submit Time'])
    time_df = time_df.rename(columns={'Case No.':'Case no.'})
    print('Finish reading data!')
    print('-'*50)
    
    #将tc/urgent/rr信息与时间信息合并
    tc = tc[(tc['Result.1']=='failed') & (tc['Urgency']=='Preventive')]
    tc = pd.merge(tc, time_df, on='Case no.', how='left')
    urgent = urgent[(urgent['Result.1']=='failed') & (urgent['Urgency']=='Urgent')]
    urgent = pd.merge(urgent, time_df, on='Case no.', how='left')
    rr = rr[(rr['Result.3']=='failed') & (rr['Urgency']=='Reply requested')]
    rr = pd.merge(rr, time_df, on='Case no.', how='left')
    time.sleep(1)
    print('Finish time integration!')
    print('-'*50)
    
    #将tc/urgent/rr中的时间信息以星期几的格式储存在Week Number列中
    #剔除tc/urgent/rr中的周末时间
    df_list = [tc, urgent, rr]
    for i in df_list:
        number = []
        for j in range(i.shape[0]):
            number.append(i.loc[j, 'Submit Time'].strftime("%w"))
        i['Week Number'] = number 
    tc = tc[(tc['Week Number'] != '6')]
    tc = tc[(tc['Week Number'] != '0')]
    urgent = urgent[(urgent['Week Number'] != '6')]
    urgent = urgent[(urgent['Week Number'] != '0')]
    rr = rr[(rr['Week Number'] != '6')]
    rr = rr[(rr['Week Number'] != '0')]
    
    #剔除之前输入的假期时间数据
    #仅保留case no.与team信息
    #更改df index
    if number_of_holiday != 0:
        for i in range(len(holiday_no_list)):
            tc = tc[(tc['Week Number'] != holiday_no_list[i])]
            urgent = urgent[(urgent['Week Number'] != holiday_no_list[i])]
            rr = rr[(rr['Week Number'] != holiday_no_list[i])]  
    tc = pd.DataFrame(tc, columns=['Case no.', 'Team'])
    tc = tc.rename(columns={'Case no.':'TC<2h'})
    tc = tc.set_index(['TC<2h'])
    urgent = pd.DataFrame(urgent, columns=['Case no.', 'Team'])
    urgent = urgent.rename(columns={'Case no.':'URGENT<2h'})
    urgent = urgent.set_index(['URGENT<2h'])
    rr = pd.DataFrame(rr, columns=['Case no.', 'Team'])
    rr = rr.rename(columns={'Case no.':'RR<4h'})
    rr = rr.set_index(['RR<4h'])
    
    #储存failed list数据
    failed_list_file_name = 'response time failed list CW%s.xlsx'%(week_number,)
    writer = pd.ExcelWriter(failed_list_file_name)
    tc.to_excel(writer)
    urgent.to_excel(writer, startcol = 2)
    rr.to_excel(writer, startcol = 4)
    
    #保存
    writer.save()
    print("Finsh failed list analyzing!")

#打印基本信息
def print_card_infors_basic_menu():
    title = "For Weekly Report, For Mother Russia!"
    option1 = "1、PuMA open/created/modified zipfiles Process"
    option2 = "2、Modified Cases for Pingpong Analyze"
    option3 = "3、Open Case from Open and Cognos Files"
    option4 = "4、Escalation Email Analyse"
    option5 = "5、Failed List from Response Time"
    option6 = "6、Other options still under programming..."
    option0 = "0、Quit"
    print("="*50)
    print(title.center(50))
    print("-"*50)
    print(option1.center(50))
    print(option2.center(50))
    print(option3.center(50))
    print(option4.center(50))
    print(option5.center(50))
    print(option6.center(50))
    print(option0.center(50))
    print("="*50)

#主函数
def main():
    while True:
        print_card_infors_basic_menu()
        flag = int(input("Input your option："))
        print('-'*50)
        if flag == 1:
            week_number = input('input week number:')
            xx(week_number)     
        elif flag == 2:
            week_number = input('input week number:')
            print('-'*50)
            print('Please be waiting for about 10.80s...')
            pp(week_number)  
        elif flag == 3:
            week_number = input("input week number:")
            end_day = input("input end day of open case(eg: 0526):")
            oc(week_number, end_day)
        elif flag == 4:
            week_number = input("input week number:")
            ee(week_number)    
        elif flag == 5:
            week_number = input("input week number:")
            print('-'*50)
            print("Make sure Response time file is completed before you analyze failed list!")
            print('-'*50)
            fl(week_number)
        elif flag == 6:
            print('This part has not finished yet...')
            time.sleep(1)
        elif flag == 0:
            break
        else:
            print('Wrong number, once again')
            time.sleep(1)

if __name__ == '__main__':
    main()