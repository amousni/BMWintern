import xlrd
from openpyxl import Workbook
import re
import datetime
import pandas as pd

#读取人名信息
def load_name_list():
    df = pd.read_csv('name_list.csv').set_index('name')
    return [list(df.index), df.T.to_dict()]

#需要匹配的人名列表，modified文件中C/F/R列title
#name_list = ['Kun Xu', 'Annamaria Zheng', 'xing bin', 'bin xing', 'Chunhu Zhang', 'Don Wang', 'Weiyu Hou', 'Kejie Lu', 'Wenjie Li', 'Hailong Zhao', 'Qi Yang', 'Yongyuan Gao', 'Ruoyun Wang', 'Frank Wang', 'Jian Xu', 'Min Song', 'Ben Ma', 'Jianquan Feng', 'Hu Sun', 'Joshua Yan', 'Raul Bernal', 'Yali Jiang', 'Kai Wang', 'Chengyan Sun', 'Hong Wang', 'Xin Tian', 'Xin Tian', 'Alice Yang', 'Fady Bawatna', 'Li Wang', 'Peter Chen', 'Angela Feng', 'Pengcheng Gao', 'lei Zhang', 'Dongli Liu', 'Lesley lan', 'Zhenhua Li', 'Alexander Lang', 'Zhi Zhang', 'Shuai Ni', 'Donghui Zhang', 'Rui Miao', 'Benzy Qin', 'Ping Hao', 'TB GZ', 'hongxing zhao', 'Rocky Chen', 'Becky Zhang', 'yuan zhou3', 'juan zhang', 'Xin Tian', 'Frank Zheng', 'Ivan Wang', 'Qiang Zhong', 'Guoxiang Deng', 'Alan Zhang', 'Test1 Puma', 'Reader Power User', 'Peter Loertzing', 'Carsten Gruetzmacher', 'Christoph Brettle', 'Tao Zhao', 'Wendy Wu', 'Baolin Xiang', 'Bin Fu', 'Joanna Zhong', 'Mahathi Murthy', 'Rishika Menon', 'Yali Jiang', 'Xiaomao Wu', 'Ting Xu', 'Yali Jiang', 'Hong Wang', 'ruiqing Zhang', 'Yang Li', 'David Li', 'Yu Hou', 'Jing Dai', 'Joanna Zhong', 'Hong Wang', 'Fengjie Lu', 'Jie Tan', 'Sean Huang', 'Dan Wu', 'Fan Bo Meng', 'Anna Yang', 'Fernando Ferrer', 'Rocky Chen', 'Becky Zhang', 'Fei Ren', 'yuan zhou3', 'juan zhang', 'Jie Yang', 'Xin Tian', 'Xin Tian', 'Jintu Yang', 'Andrew Fletcher', 'Yunqian Wei', 'Xiaoxi Liu', 'Alice Yang', 'Bo Yang', 'Fady Bawatna', 'Huifeng Cong', 'Li Wang', 'Peter Chen', 'Marshal You', 'Zuoqi Qu', 'Graham Boyd', 'Jamie Zhang', 'Angela Feng', 'Yanjun Gu', 'Pengcheng Gao', 'Fangchao Xu', 'Pengcheng Gao', 'Yan Yang', 'Lei Li', 'Leo Li', 'Juan Zhang', 'Shixin Dong', 'juan zhang', 'Zhaojun Li', 'Jie Tan', 'David Wang', 'Marc Ma', 'Dawei Wang', 'Dongsheng Liu ', 'Zhengyun Lu', 'Lipeng Wang', 'Lei Liu', 'Yuping li', 'lei Zhang', 'Quan Huang', 'Christian Schoenmeyer', 'Maple Huo', 'Laney Zhang', 'Joachim Fluegge', 'Gordon Zhao', 'Chris Dong', 'Artem Lobanow', 'Zhe Yang', 'Gary Tian', 'Zhicong Xie', 'Dongli Liu', 'Lesley lan', 'Steven Miao', 'Alex Li', 'Zhenhua Li', 'Hao Ling', 'Ao Shen', 'Frank Brandel', 'Mareike Volkmer', 'Arpad Toth', 'Xuqiang Yin', 'Qiang Zhong', 'Andreas Sedlak', 'Ting Jin', 'Zhihai Wu', 'Chun Zou', 'Yuan Zhang', 'Collie Wang', 'Yiding Zhang', 'Guoxiang Deng', 'Season Liu', 'Jonathan Zhu', 'Billy Xiao', 'Yin Li', 'Alan Zhang', 'Ke Chen', 'Justin Jiang', 'Alexander Lang', 'Weiming Hou', 'Gabrielle Ge', 'Billy Fang', 'Wilson Xu', 'Dawei Zhang', 'Zhi Zhang', 'Shuai Ni', 'Joao Martins', 'Wei Shi', 'Gavin Cai', 'dan zhang', 'Justin Zeng', 'Jimmy Ji', 'Spike Zhao', 'Juliet Zhang', 'Jiangtao Chi', 'Steven Jiang', 'Weiliang Huang', 'Ruifeng Liu', 'Ming Liu', 'Yan Ke', 'Vicente Liu', 'Chris Liu', 'Mandy Du', 'Fred Zhang', 'Qinran Luo', 'Harry Zhang', 'Ming Yuan', 'Benzy Qin', 'Jian Ma', 'Walter Gassner', 'Steven Wu', 'Galen Guo', 'Feal Wang', 'Colin Song', 'Fred Feng', 'Ping Hao', 'Jianmin Xin', 'Jarod Yang', 'TB GZ', 'Tony Xu', 'Mingjia Ma', 'Tianshu Li', 'Rich Fu', 'Wayne Zhang', 'Alexander Lang', 'hongxing zhao', 'Christian Kluth', 'Friedrich Mursch', 'Alexander Fischer', 'John White', 'Richard Brown', 'Test1 Puma', 'Reader Power User', 'Thomas Riedl', 'Mark Werner', 'Harald Kammerl', 'Jean Claude Chauveau', 'Frank Schoenrock', 'Stefan Rosse', 'Horst Redner', 'Regan Chen', 'Friedrich Mursch', 'Ian Weston', 'Andreas Herkert', 'Sebastien Beck', 'Thomas Drexler', 'Markus Gamisch', 'Wenjin Chen', 'Richard Wimmer', 'Huashan Wang', 'Yingxue Peng', 'Stefan Hofmann', 'Bob White', 'Christian Mueck', 'Wendi Yi', 'Feng Fred', 'Di Lai', 'Armin Meyer', 'Maoqi Cui', 'Hebe He', 'zhihai wu', 'Sinan Oeztan', 'Yingxue Peng', 'Rex Jiang', 'Frank Zintl', 'Sherry Lu', 'Junli Liang', 'Tim Shi', 'Yuhuo Ma', 'lilong zhao', 'Allen Mao', 'Hailong Li', 'Shawn Chen', 'Leiming He', 'Ermera Setyadi', 'Kimreol Chen', 'Stefan Krott', 'Xiaohui Li', 'Liping Hong', 'Jay Lee', 'Christian Klenovsky', 'Feixue Li', 'Jeffrey Jiang', 'Shelina Zhao', 'Zhang Zhe', 'Jian Guo', 'weiqiang mu', 'Sabrina Fischer', 'Zhi Cui', 'Ryan Qian', 'Zenan Wen', 'Yinbo Shen', 'Zhaoxian Chu', 'Dragon You', 'Wenhao Li', 'Zutao Kou', 'Lunan Tan', 'Charlie Geng', 'Sandy Lu', 'Ke Zhou', 'Haoxin Liu', 'Gang Li', 'Mingjun Jin', 'Levi Yang', 'Junyan Liu', 'Yu Wang', 'Zhiqiang Chen', 'Yongcheng Shu', 'Shengnan Si', 'Lei Yang ', 'Bin Xing', 'Hanan Wu', 'Jack Zhang']

modified_index = ['Case no.', 'VIN no. (last 7 digits)', 'Previous recommendations/queries/additional information']

#匹配recommendation与additional information中的回复人信息与回复次数
#需要注意的是：
#   1.additional information中会有一部分回复信息
#   2.recommendation与additional information中有回复时间，仅统计半年内的回复信息
#   3.recommendation的回复人一定是组内的人，additional information中不一定
def recom_pair(recom_str):
    
    unfinished_name_list = load_name_list()[0]
    name_list = []
    for i in unfinished_name_list:
        last_first = i.split(' ')
        name = last_first[-1] + ' ' + last_first[0]
        name_list.append(name)
    
    #时间节点，当天至半年前
    end_day = datetime.date.today()
    start_day = end_day - datetime.timedelta(days=183)
    
    #正则匹配回复时间与回复者人名
    pattern = re.compile('Recommendation   (.*?)   .*?   (.*? .*?) ')
    
    #匹配信息，一个元组列表，元组中第一个元素为回复时间，第二个元素为回复人名
    s = pattern.findall(recom_str)
    
    #人名列表
    recom_name_list = []
    
    #时间匹配计数
    flag = 0
    for i in s:
        mdy_str = i[0].split('/')
        mdy = datetime.date(int(str(20) + mdy_str[2]),int(mdy_str[0]),int(mdy_str[1]))
        if mdy <= end_day and mdy >= start_day:
            recom_name_list.append(i[1])
            flag += 1
            
    #Additional information正则匹配
    pattern_a = re.compile('Additional information   (.*?)   .*?   (.*? .*?) ')
    s_a = pattern_a.findall(recom_str)
    for i in s_a:
        mdy_str = i[0].split('/')
        mdy = datetime.date(int(str(20) + mdy_str[2]),int(mdy_str[0]),int(mdy_str[1]))
        if mdy <= end_day and mdy >= start_day:
            if i[1] in name_list:
                recom_name_list.append(i[1])
                flag += 1
                
    #人名合并
    name = ''
    recom_name = set(recom_name_list)
    recom_name_list = list(recom_name)
    for i in range(len(recom_name_list)):
        if i == 0:
            name = name + recom_name_list[i]
        else:
            name = name + '|' + recom_name_list[i]
            
    #返回匹配个数与人名信息
    return [flag, name]
        
#Pingpong分析
def modified_4_pingpong(week_number):

    #读取本周modified文件
    #week_number为本周周报号, file_name表示本周modified文件名
    file_name = 'CW' + week_number + ' MODIFIED' + '.xlsx'

	#wb为本周modified文件
    wb = xlrd.open_workbook(filename = file_name)

	#打开wb中sheet1
    sheet1 = wb.sheet_by_index(0)
    cla = sheet1.row_values(0)

	#sheet1中的C/F/R列
    case_no = cla.index(modified_index[0])
    vin = cla.index(modified_index[1])
    recom = cla.index(modified_index[2])
    #sheet1中列存为列表方便写入新文件
    sheet_index_list = [case_no, vin, recom]
    data_number = len(sheet1.col_values(case_no))

    #记录数据
    #case_no_value = sheet1.col_values(case_no)[0:]
    #vin_value = sheet1.col_values(vin)[0:]
    recom_value = sheet1.col_values(recom)[0:]

    #将C/F/R列写入新文件
    pingpong_file_name = 'Ping-Pong Analyzing CW' + week_number + '.xlsx'
    wb_pingpong = Workbook()
    sheet = wb_pingpong.active
    sheet.title = 'Statisitc'
    #三列，储存在sheet_index_list中
    #flag读写行数标注
    #flag = 1
    for row in range(data_number):
        #第1行记录表头信息
        if row == 0:
            sheet.cell(row+1,1,sheet1.col_values(sheet_index_list[0])[row])
            sheet.cell(row+1,2,sheet1.col_values(sheet_index_list[1])[row])
            sheet.cell(row+1,5,sheet1.col_values(sheet_index_list[2])[row])
            sheet.cell(row+1,4,'Recommend Person')
            sheet.cell(row+1,3,'Numbers OF reply')
        else:
            #记录前三列C/F/R列信息
            sheet.cell(row+1,1,sheet1.col_values(sheet_index_list[0])[row])
            sheet.cell(row+1,2,sheet1.col_values(sheet_index_list[1])[row])
            sheet.cell(row+1,5,sheet1.col_values(sheet_index_list[2])[row])
            #recommendation_string
            recom_str = recom_value[row]                
            pair_info = recom_pair(recom_str)
            sheet.cell(row+1,4,pair_info[1])
            sheet.cell(row+1,3,pair_info[0])

    #保存Pingpong文件，xlsx格式保存
    wb_pingpong.save(pingpong_file_name)
    print('-'*10)
    print('Finish Pingpong analyzing!')

def main():
    week_number = input("Input week number:")
    modified_4_pingpong(week_number)

if __name__ == '__main__':
    main()